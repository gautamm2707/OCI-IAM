import openpyxl as xl
workbook = xl.load_workbook(filename="/Users/gautamm/Downloads/HelpNow-Data.xlsx")
ws = workbook.active
word1 = "client id"
word2 = "clientid"
word3 = "client_id"
word4 = "client_id:"


dum = "whaere is the client id is here"
rows = ws.max_row

for i in range(1, rows):
    st = ws.cell(row=i, column=3).value
    #print(type(st))
    #print(st)
    if st is None :
        #print ("caught")
        continue
    # print(type(st))
    st = st.lower()
    #print(st)
    #if word1 in st or word2 in st or word3 in st or word4 in st:
    if word5 in st:
        temp = ws.cell(row=i, column=2).value
        print(temp)

