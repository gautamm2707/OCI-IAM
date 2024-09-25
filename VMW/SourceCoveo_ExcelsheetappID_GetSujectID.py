
import json
import requests
import base64
import openpyxl as xl
import urllib3
from zipfile import ZipFile
urllib3.disable_warnings()

#workbook1 = xl.load_workbook(filename="/Users/gautamm/Downloads/WS1_Selected_Apps.xlsx")
workbook1 = xl.load_workbook(filename="/Users/gautamm/Desktop/test.xlsx")
ws1 = workbook1.active
rows1 = ws1.max_row

def get_encoded(clid,clsecret):
    encoded = clid + ":" + clsecret
    baseencoded = base64.urlsafe_b64encode(encoded.encode('UTF-8')).decode('ascii')
    #print("Base64encoded string:", baseencoded)
    return baseencoded

def get_access_token(url,header):
    parameter = {'grant_type': "client_credentials"}
    response = requests.post(url, headers=header, params=parameter, verify=False)
    jsonresp = json.loads(response.content)
    access_token = jsonresp.get('access_token')
    return access_token

api_urlbase = "https://###########.com/SAAS/"
clid = "##########"
clsecret = "###############"
encodedtoken = get_encoded(clid, clsecret)
extra = "auth/oauthtoken"
headers = {'Content-Type' : 'application/x-www-form-urlencoded' , 'Authorization' : 'Basic %s' % encodedtoken}
accesstoken = get_access_token(api_urlbase+extra, headers)
#print(accesstoken)

searchurl = "https://##############.com/SAAS/jersey/manager/api/catalogitems/search?startIndex=0&pageSize=1500"
headers2 = {'Authorization': 'Bearer ' + accesstoken, 'Accept': 'application/vnd.vmware.horizon.manager.catalog.item.list+json' , 'Content-Type': 'application/vnd.vmware.horizon.manager.catalog.search+json'}
print(rows1)
#Getting Apps
mainapplist = []
for i in range(1, rows1+1):
    appname = ws1.cell(row=i, column=2).value
    uuid = ws1.cell(row=i, column=1).value
    data = '{"nameFilter": "'+appname+'" ,"includeTypes":["Saml11","Saml20","WSFed12","WebAppLink", "AnyApp"], "categories":[], "rootResource":"false" }'
    #print(uuid)
    #print(appname)
    resp = requests.post(searchurl, headers=headers2, verify=False, data=data)
    jsonresp = json.loads(resp.content)
    tempjsn = jsonresp.get("items")
    size = jsonresp.get("totalSize")

    if size == 0:
        print(appname + " : is not present in WS1")
        continue
    #print(tempjsn)


    #trim the json
    trimjson = {}
    trimjson["uuid"] = tempjsn[0].get("uuid")
    name = trimjson["name"] = tempjsn[0].get("name")
    trimjson["description"] = tempjsn[0].get("description")
    trimjson["catalogItemType"] = tempjsn[0].get("catalogItemType")
    trimjson["labels"] = tempjsn[0].get("labels")
    trimjson["visible"] = tempjsn[0].get("visible")
    trimjson["internal"] = tempjsn[0].get("internal")
    trimjson["_links"] = tempjsn[0].get("_links")
    trimjson["items"] = tempjsn[0].get("items")
    #print(trimjson)
    print(name)

    u = "https://############.com/SAAS/jersey/manager/api/entitlements/definitions/catalogitems/" + uuid
    headers3 = {'Authorization': 'Bearer ' + accesstoken,
                'Accept': 'application/vnd.vmware.horizon.manager.entitlements.v2.definition.list+json',
                'Content-Type': 'application/vnd.vmware.horizon.manager.entitlements.v2.definition.list+json'}
    resp1 = requests.get(u, headers=headers3, verify=False)
    jsonresp1 = json.loads(resp1.content)
    mainapplist.append(trimjson)
    #mainapplist.append(uuid)
    mainapplist.append(jsonresp1)

print(mainapplist)

with open("app.json", "w") as outfile:
    json.dump(mainapplist, outfile)
with ZipFile('app.zip', 'w') as zipp:
    zipp.write("app.json")





