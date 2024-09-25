import json
import requests
import base64
import urllib3
import openpyxl as xl
urllib3.disable_warnings()
new_wb = xl.Workbook()
sheet = new_wb.active

#Base64encoding
def get_encoded(clid,clsecret):
    encoded = clid + ":" + clsecret
    baseencoded = base64.urlsafe_b64encode(encoded.encode('UTF-8')).decode('ascii')
    #print("Base64encoded string:", baseencoded)
    return baseencoded

#GeneratingAccessToken
def get_access_token(url,header):
    parameter = {'grant_type': "client_credentials"}
    response = requests.post(url, headers=header, params=parameter, verify=False)
    jsonresp = json.loads(response.content)
    access_token = jsonresp.get('access_token')
    return access_token

#main
api_urlbase = "https://######.########.com/SAAS/"
clid = "##########"
clsecret = "###########"
encodedtoken = get_encoded(clid, clsecret)
extra = "auth/oauthtoken"
headers = {'Content-Type' : 'application/x-www-form-urlencoded' , 'Authorization' : 'Basic %s' % encodedtoken}
accesstoken = get_access_token(api_urlbase+extra, headers)
print(accesstoken)

#Getting Apps ID
searchurl = "https://######.########.com/SAAS/jersey/manager/api/catalogitems/search?startIndex=0&pageSize=2000"
headers2 = {'Authorization': 'Bearer ' + accesstoken, 'Accept': 'application/vnd.vmware.horizon.manager.catalog.item.list+json' , 'Content-Type': 'application/vnd.vmware.horizon.manager.catalog.search+json'}
data = '{"includeTypes":["Saml11","Saml20","WSFed12","WebAppLink", "AnyApp"], "categories":[], "rootResource":"false" }'
resp = requests.post(searchurl, headers=headers2, verify=False, data=data)

jsonresp = json.loads(resp.content)
tempjsn = jsonresp.get("items")
size = jsonresp.get("totalSize")
#print(tempjsn)
#print(size)
mainapplist = []
mainsubjectidlist =[]
c = 2
for i in range(3):
    appId = tempjsn[i].get("uuid")
    if (tempjsn[i].get("visible") == True):

        trimjson = {}
        name = trimjson["name"] = tempjsn[i].get("name")
        authinfo = tempjsn[i].get("authInfo")
        attributes = authinfo.get("attributes")
        sheet.cell(row=i+1, column=1).value = name
        print(name)
        if(attributes is None):
            print("")
        else:
            for attribute in attributes:
                sheet.cell(row=i+1, column=c).value = attribute.get("name")
                c = +1
                print(attribute.get("name"))

new_wb.save(filename="/Users/gautamm/Downloads/AppAttributes.xlsx")




