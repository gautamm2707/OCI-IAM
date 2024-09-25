import openpyxl as xl
workbook = xl.load_workbook(filename="/Users/gautamm/Downloads/HelpNow-Data.xlsx")
new_wb = xl.Workbook()
sheet = new_wb.active
ws = workbook.active
global temp
"""word1 = "client id"
word2 = "clientid"
word3 = "client_id"
word4 = "client_id:"""""
word5 = "client"

dum = "whaere is the client id is here"
rows = ws.max_row
j = 1
for i in range(1, rows):
    st = ws.cell(row=i, column=3).value
    #print(type(st))
    #print(st)
    if st is None :
        #print ("caught")
        continue
    # print(type(st))
    st = st.lower()
    #print(st)
    #if word1 in st or word2 in st or word3 in st or word4 in st:
    global temp
    if word5 in st:
        temp = ws.cell(row=i, column=2).value
        sheet.cell(row=j , column=1).value = temp
        sheet.cell(row=j , column=2).value = st
        j += 1
        print(j)

new_wb.save(filename="/Users/gautamm/Downloads/NewExcel.xlsx")



