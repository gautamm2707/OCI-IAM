import openpyxl as xl
workbook1 = xl.load_workbook(filename="/Users/gautamm/Downloads/HelpNow-Data.xlsx")
workbook2 = xl.load_workbook(filename="/Users/gautamm/Downloads/ClientList-Stg.xlsx")
new_wb = xl.Workbook()
sheet = new_wb.active
ws1 = workbook1.active
ws2 = workbook2.active
global temp
word1 = "client id"
word2 = "clientid"
word3 = "client_id"
word4 = "client_id:"
word5 = "client"

dum = "whaere is the client id is here"
rows1 = ws1.max_row
rows2 = ws2.max_row
k = 1
for j in range(1, rows2):
    for i in range(1, rows1):
        cl = ws2.cell(row=j, column=1).value
        st = ws1.cell(row=i, column=3).value

        if st is None :
            continue
        st = st.lower()
        if cl is None :
            continue
        cl = cl.lower()

        global temp
        if cl in st and (word1 in st or word2 in st or word3 in st or word3 in st or word4 in st ) :
            temp = ws1.cell(row=i, column=2).value
            sheet.cell(row=k, column=1).value = temp
            sheet.cell(row=k, column=2).value = st
            k += 1
            print(k)

new_wb.save(filename="/Users/gautamm/Downloads/NewExcel.xlsx")



