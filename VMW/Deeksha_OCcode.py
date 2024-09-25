import json
import requests
import base64
import openpyxl as xl
import urllib3
urllib3.disable_warnings()

workbook1 = xl.load_workbook(filename="/Users/gautamm/Downloads/WS1_Selected_Apps.xlsx")
ws1 = workbook1.active
rows1 = ws1.max_row

def get_encoded(clid,clsecret):
    encoded = clid + ":" + clsecret
    baseencoded = base64.urlsafe_b64encode(encoded.encode('UTF-8')).decode('ascii')
    #print("Base64encoded string:", baseencoded)
    return baseencoded

def get_access_token(url,header):
    parameter = {'grant_type': "client_credentials"}
    response = requests.post(url, headers=header, params=parameter, verify=False)
    jsonresp = json.loads(response.content)
    access_token = jsonresp.get('access_token')
    return access_token

api_urlbase = "https://#######.#########.com/SAAS/"
clid = "###########"
clsecret = "##############"
encodedtoken = get_encoded(clid, clsecret)
extra = "auth/oauthtoken"
headers = {'Content-Type' : 'application/x-www-form-urlencoded' , 'Authorization' : 'Basic %s' % encodedtoken}
accesstoken = get_access_token(api_urlbase+extra, headers)
print(accesstoken)


#Search apps
# f=open("test.txt", "r")
# dec = f.read().splitlines()
searchurl = "https://##############.com/SAAS/jersey/manager/api/catalogitems/search?startIndex=0&pageSize=15000"
headers2 = {'Content-Type' : 'application/vnd.vmware.horizon.manager.catalog.search+json' , 'Accept' : 'application/vnd.vmware.horizon.manager.catalog.item.list+json', 'Authorization' : 'Bearer ' + accesstoken}

#print(headers2)

mainlist = []
for i in range(1,rows1):
    cl = ws1.cell(row=i, column=2).value

    reqbody = '{ "nameFilter": "'+cl+'", "includeTypes":["Saml20"], "categories":["saas"], "includeAttributes":["labels"], "includeIconBytes":"true" }'
    resp = requests.post(searchurl, headers=headers2, data=reqbody, verify=False)
    print(resp.content)
    jsonresp = json.loads(resp.content)
    tempjsn = jsonresp.get("items")
    sizett = jsonresp.get("totalSize")
    #print(sizett)
    if sizett == 0:
        print("Caught")
        continue
    dictofids = {}
    dictofids["uuid"] = tempjsn[0].get("uuid")
    dictofids["orgname"] = i
    dictofids["ws1name"] = tempjsn[0].get("name")
    #print(dictofids["ws1name"])
    mainlist.append(dictofids)
print(mainlist)


#Get app information, modify and update app
"""for nm in mainlist:
    uuid = nm.get("uuid")
    orgname = nm.get("orgname")
    requrl = "https://######.#######.com/SAAS/jersey/manager/api/catalogitems/" + uuid
    print(requrl)
    headers3 = {'Accept' : 'application/vnd.vmware.horizon.manager.catalog.saml20+json', 'Authorization' : 'Bearer ' + accesstoken}
    neworgurl = "https://#######.oc.#####.com/tenant/"+orgname+"/"
    print(neworgurl)
    resp1 = requests.get(requrl, headers=headers3)
    json_object = json.loads(resp1.content)
    json_object["authInfo"]["loginRedirectionUrl"] = neworgurl
    #update app
    headers4 = {'Content-Type' : 'application/vnd.vmware.horizon.manager.catalog.saml20+json', 'Accept': 'application/vnd.vmware.horizon.manager.catalog.saml20+json', 'Authorization': 'Bearer ' + accesstoken}
    updatedbody = json.dumps(json_object)
    #print(updatedbody)
    newresp = requests.put(requrl, headers=headers4, data=updatedbody)
    print(newresp)"""
