
import json
import requests
import base64
import openpyxl as xl
import urllib3
urllib3.disable_warnings()

workbook1 = xl.load_workbook(filename="/Users/gautamm/Downloads/WS1_Selected_Apps.xlsx")
ws1 = workbook1.active
rows1 = ws1.max_row

def get_encoded(clid,clsecret):
    encoded = clid + ":" + clsecret
    baseencoded = base64.urlsafe_b64encode(encoded.encode('UTF-8')).decode('ascii')
    #print("Base64encoded string:", baseencoded)
    return baseencoded

def get_access_token(url,header):
    parameter = {'grant_type': "client_credentials"}
    response = requests.post(url, headers=header, params=parameter, verify=False)
    jsonresp = json.loads(response.content)
    access_token = jsonresp.get('access_token')
    return access_token

api_urlbase = "https://myvmware.workspaceair.com/SAAS/"
clid = "gm_cli"
clsecret = "5FwnIOUlTRle0KjwuIizJvqwJPuJHfUP96IbpQWb19w1oW8k"
encodedtoken = get_encoded(clid, clsecret)
extra = "auth/oauthtoken"
headers = {'Content-Type' : 'application/x-www-form-urlencoded' , 'Authorization' : 'Basic %s' % encodedtoken}
accesstoken = get_access_token(api_urlbase+extra, headers)
print(accesstoken)


searchurl = "https://myvmware.workspaceair.com/SAAS/jersey/manager/api/reporting/reports/appentitlement"
headers2 = {'Authorization' : 'Bearer ' + accesstoken}

#For Testing with Test Application
"""param = {'appId': "2e349a46-9ea7-4181-bac8-3d2cba54b54d"}
resp = requests.get(searchurl, headers=headers2, verify=False, params=param)
jsonresp = json.loads(resp.content)
    #tempjsn = jsonresp.get("data")
tempjsn = jsonresp.get("data")
for row in tempjsn:
    if row == 0:
        print("Caught")
        continue
    print(row[1])
"""


for i in range(1, rows1):
    uuid = ws1.cell(row=i, column=1).value
    param = {'appId': uuid}
    print(uuid)
    resp = requests.get(searchurl, headers=headers2, verify=False, params=param)
    jsonresp = json.loads(resp.content)
    tempjsn = jsonresp.get("data")
    print("successful")

"""    for row in tempjsn:
        print(row[1])

"""
